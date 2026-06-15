"""Excel (.xlsx / .xls) のセル値検索。

.xlsx は openpyxl read_only。本走査の前に sharedStrings.xml を
zip 直読みして「文字列キーワードが存在しないブックを即スキップ」する
高速プレフィルタを備える。.xls は xlrd で読む。
"""
from __future__ import annotations

import re
import zipfile
from collections.abc import Iterator
from pathlib import Path

# openpyxl / xlrd は重い（起動時 ~200ms）ため、Excel を実際に検索する時だけ
# 遅延ロードする（起動・Win+E の体感を速くする）。
_openpyxl = None
_xlrd = None
_get_col = None


def _load_openpyxl():
    global _openpyxl, _get_col
    if _openpyxl is None:
        import openpyxl
        from openpyxl.utils import get_column_letter
        _openpyxl = openpyxl
        _get_col = get_column_letter
    return _openpyxl, _get_col


def _load_xlrd():
    global _xlrd, _get_col
    if _xlrd is None:
        import xlrd
        from openpyxl.utils import get_column_letter
        _xlrd = xlrd
        _get_col = get_column_letter
    return _xlrd, _get_col


_NUMERIC_RE = re.compile(r"^[\d.,\-+eE]+$")
_TAG_RE = re.compile(rb"<[^>]+>")


def may_contain_keyword(filepath: str | Path, keyword: str) -> bool:
    """sharedStrings.xml の高速プレフィルタ。

    False なら文字列セルにキーワードは確実に存在しない（本走査不要）。
    数値系キーワードは数値セルに入り得る（sharedStrings に載らない）ため
    常に True。判定不能（破損等）も True を返し本走査に委ねる。
    """
    if _NUMERIC_RE.match(keyword):
        return True
    needle = keyword.lower()
    try:
        with zipfile.ZipFile(filepath) as zf:
            # 共有文字列に加え、インライン文字列（openpyxl 出力等）も
            # 持ち得るワークシートXMLを生のまま走査する
            targets = [n for n in zf.namelist()
                       if n == "xl/sharedStrings.xml"
                       or (n.startswith("xl/worksheets/")
                           and n.endswith(".xml"))]
            for name in targets:
                data = zf.read(name)
                text = _TAG_RE.sub(b"", data).decode(
                    "utf-8", errors="replace")
                if needle in text.lower():
                    return True
            return False
    except (OSError, zipfile.BadZipFile, KeyError):
        return True  # 判定不能は本走査に委ねる


def search_in_excel(filepath: str | Path, keyword: str,
                    case_sensitive: bool = False,
                    max_hits: int = 100) -> Iterator[tuple[str, str, str]]:
    """全シートを走査し (シート名, セル番地, セル値) を逐次返す。

    read_only モードで開くため大きなブックでもメモリを食わない。
    """
    needle = keyword if case_sensitive else keyword.lower()
    if not may_contain_keyword(filepath, keyword):
        return
    openpyxl, get_column_letter = _load_openpyxl()
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True,
                                    data_only=True)
    except Exception:  # 破損ブック・暗号化等は対象外としてスキップ
        return
    hits = 0
    try:
        for ws in wb.worksheets:
            for row_idx, row in enumerate(ws.iter_rows(values_only=True),
                                          start=1):
                for col_idx, value in enumerate(row, start=1):
                    if value is None:
                        continue
                    text = str(value)
                    haystack = text if case_sensitive else text.lower()
                    if needle in haystack:
                        address = f"{get_column_letter(col_idx)}{row_idx}"
                        yield ws.title, address, text[:200]
                        hits += 1
                        if hits >= max_hits:
                            return
    finally:
        wb.close()


def _xls_cell_text(cell) -> str | None:
    xlrd, _ = _load_xlrd()
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    value = cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER and value == int(value):
        value = int(value)  # 1.0 → 1 表記
    return str(value)


def search_in_xls(filepath: str | Path, keyword: str,
                  case_sensitive: bool = False,
                  max_hits: int = 100) -> Iterator[tuple[str, str, str]]:
    """.xls（旧形式）のセル値検索。xlrd 使用。"""
    needle = keyword if case_sensitive else keyword.lower()
    xlrd, get_column_letter = _load_xlrd()
    try:
        book = xlrd.open_workbook(str(filepath), on_demand=True)
    except Exception:  # 破損・暗号化・非xlsはスキップ
        return
    hits = 0
    try:
        for sheet in book.sheets():
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    text = _xls_cell_text(sheet.cell(row_idx, col_idx))
                    if text is None:
                        continue
                    haystack = text if case_sensitive else text.lower()
                    if needle in haystack:
                        address = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                        yield sheet.name, address, text[:200]
                        hits += 1
                        if hits >= max_hits:
                            return
    finally:
        book.release_resources()
