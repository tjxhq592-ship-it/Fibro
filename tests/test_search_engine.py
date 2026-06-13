import threading

import openpyxl
import pytest

from app.engine.search_engine import (
    SearchHit, SearchMode, SearchOptions, SearchStats, search,
)
from app.engine.text_reader import detect_encoding, is_binary, search_in_text
from app.engine.excel_reader import search_in_excel


@pytest.fixture
def tree(tmp_path):
    """検索用のサンプルツリーを作る。"""
    (tmp_path / "report_2026.txt").write_text(
        "first line\nthe ORDER_ID is here\nlast line", encoding="utf-8")
    (tmp_path / "notes.md").write_text("nothing relevant", encoding="utf-8")
    (tmp_path / "binary.bin").write_bytes(b"\x00\x01\x02order_id\x00")
    (tmp_path / "japanese.txt").write_text(
        "重要なメモ: order_id を確認", encoding="cp932")
    sub = tmp_path / "sub"
    sub.mkdir()
    (sub / "order_id_list.csv").write_text("a,b,c", encoding="utf-8")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["C5"] = "order_id"
    ws["A1"] = "unrelated"
    wb.save(tmp_path / "book.xlsx")
    return tmp_path


class TestTextReader:
    def test_binary_detection(self, tree):
        assert is_binary(tree / "binary.bin")
        assert not is_binary(tree / "report_2026.txt")

    def test_encoding_fallback_cp932(self, tree):
        assert detect_encoding(tree / "japanese.txt") == "cp932"

    def test_search_case_insensitive(self, tree):
        hits = list(search_in_text(tree / "report_2026.txt", "order_id"))
        assert hits == [(2, "the ORDER_ID is here")]

    def test_search_cp932_content(self, tree):
        hits = list(search_in_text(tree / "japanese.txt", "重要"))
        assert len(hits) == 1


class TestExcelReader:
    def test_cell_hit(self, tree):
        hits = list(search_in_excel(tree / "book.xlsx", "order_id"))
        assert hits == [("Data", "C5", "order_id")]

    def test_corrupt_file_skipped(self, tmp_path):
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not a zip")
        assert list(search_in_excel(bad, "x")) == []


class TestFilenameSearch:
    def test_recursive_filename(self, tree):
        opts = SearchOptions(keyword="order_id",
                             modes={SearchMode.FILENAME})
        hits = list(search(tree, opts))
        assert [h.path for h in hits] == [str(tree / "sub" / "order_id_list.csv")]

    def test_non_recursive(self, tree):
        opts = SearchOptions(keyword="order_id",
                             modes={SearchMode.FILENAME}, recursive=False)
        assert list(search(tree, opts)) == []


class TestContentSearch:
    def test_text_mode(self, tree):
        opts = SearchOptions(keyword="order_id", modes={SearchMode.TEXT})
        hits = list(search(tree, opts))
        paths = {h.path for h in hits}
        assert str(tree / "report_2026.txt") in paths
        assert str(tree / "japanese.txt") in paths
        # バイナリと非対象拡張子は含まれない
        assert str(tree / "binary.bin") not in paths

    def test_excel_mode(self, tree):
        opts = SearchOptions(keyword="order_id", modes={SearchMode.EXCEL})
        hits = list(search(tree, opts))
        assert len(hits) == 1
        assert hits[0].detail == "Data!C5: order_id"

    def test_combined_modes(self, tree):
        opts = SearchOptions(
            keyword="order_id",
            modes={SearchMode.FILENAME, SearchMode.TEXT, SearchMode.EXCEL})
        kinds = {h.kind for h in search(tree, opts)}
        assert kinds == {SearchMode.FILENAME, SearchMode.TEXT,
                         SearchMode.EXCEL}

    def test_size_limit_skips(self, tree):
        opts = SearchOptions(keyword="order_id", modes={SearchMode.TEXT},
                             max_file_size=10)
        stats = SearchStats()
        hits = list(search(tree, opts, stats=stats))
        assert hits == []
        assert stats.skipped > 0

    def test_ext_filter(self, tree):
        opts = SearchOptions(keyword="order_id", modes={SearchMode.TEXT},
                             extensions={".md"})
        assert list(search(tree, opts)) == []


class TestCancel:
    def test_cancel_stops_immediately(self, tree):
        cancel = threading.Event()
        cancel.set()
        opts = SearchOptions(keyword="order_id",
                             modes={SearchMode.FILENAME})
        assert list(search(tree, opts, cancel=cancel)) == []

    def test_cancel_mid_stream(self, tmp_path):
        for i in range(100):
            (tmp_path / f"match_{i}.txt").write_text("x")
        cancel = threading.Event()
        opts = SearchOptions(keyword="match", modes={SearchMode.FILENAME})
        results: list[SearchHit] = []
        for hit in search(tmp_path, opts, cancel=cancel):
            results.append(hit)
            if len(results) == 5:
                cancel.set()
        assert len(results) <= 6  # キャンセル後すぐ停止

    def test_empty_keyword(self, tree):
        opts = SearchOptions(keyword="")
        assert list(search(tree, opts)) == []
