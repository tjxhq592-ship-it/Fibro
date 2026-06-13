"""FTS5インデックス・sharedStringsプレフィルタ・.xls対応のテスト。"""
import threading

import openpyxl
import pytest
import xlwt

from app.engine.excel_reader import (
    may_contain_keyword, search_in_excel, search_in_xls,
)
from app.engine.index_engine import SearchIndex
from app.engine.search_engine import SearchMode, SearchOptions, search


# ---- SQLite FTS5 インデックス ----

@pytest.fixture
def tree(tmp_path):
    data = tmp_path / "data"  # インデックスDBと分離
    data.mkdir()
    (data / "report_order.txt").write_text("x")
    (data / "日本語レポート.md").write_text("x")
    sub = data / "sub"
    sub.mkdir()
    (sub / "order_list.csv").write_text("x")
    return data


class TestSearchIndex:
    def test_build_and_query(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "ix" / "test.db")
        count = idx.build(tree)
        assert count == 3
        hits = idx.query(tree, "order")
        assert sorted(hits) == sorted([
            str(tree / "report_order.txt"), str(tree / "sub" / "order_list.csv")])
        idx.close()

    def test_japanese_trigram(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "test.db")
        idx.build(tree)
        assert idx.query(tree, "レポート") == [str(tree / "日本語レポート.md")]
        idx.close()

    def test_short_keyword_like_fallback(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "test.db")
        idx.build(tree)
        assert len(idx.query(tree, "or")) == 2  # 2文字 → LIKE
        idx.close()

    def test_case_insensitive(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "test.db")
        idx.build(tree)
        assert len(idx.query(tree, "ORDER")) == 2
        idx.close()

    def test_rebuild_replaces(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "test.db")
        idx.build(tree)
        (tree / "report_order.txt").unlink()
        idx.build(tree)
        assert idx.query(tree, "order") == [str(tree / "sub" / "order_list.csv")]
        idx.close()

    def test_indexed_at(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "test.db")
        assert idx.indexed_at(tree) is None
        idx.build(tree)
        assert idx.indexed_at(tree) is not None
        idx.close()

    def test_cancel_returns_minus_one(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "test.db")
        cancel = threading.Event()
        cancel.set()
        assert idx.build(tree, cancel=cancel) == -1
        idx.close()

    def test_fts_operators_neutralized(self, tree, tmp_path):
        idx = SearchIndex(tmp_path / "test.db")
        idx.build(tree)
        # FTS5 演算子を含むキーワードでもエラーにならない
        assert idx.query(tree, 'a"b OR c*') == []
        idx.close()


# ---- sharedStrings プレフィルタ ----

@pytest.fixture
def xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "order_id"
    ws["B2"] = 12345
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


class TestSharedStringsPrefilter:
    def test_present_keyword_passes(self, xlsx):
        assert may_contain_keyword(xlsx, "order_id")

    def test_absent_keyword_filtered(self, xlsx):
        assert not may_contain_keyword(xlsx, "zzz_not_here")

    def test_numeric_keyword_always_passes(self, xlsx):
        # 数値はsharedStringsに載らないため常に本走査へ
        assert may_contain_keyword(xlsx, "12345")

    def test_search_still_finds_numeric_cell(self, xlsx):
        hits = list(search_in_excel(xlsx, "12345"))
        assert hits == [("Sheet", "B2", "12345")]

    def test_search_skips_fast_when_absent(self, xlsx):
        assert list(search_in_excel(xlsx, "zzz_not_here")) == []


# ---- .xls 対応 ----

@pytest.fixture
def xls(tmp_path):
    wb = xlwt.Workbook()
    ws = wb.add_sheet("古いシート")
    ws.write(4, 2, "order_id")     # C5
    ws.write(0, 0, 99)
    path = tmp_path / "legacy.xls"
    wb.save(str(path))
    return path


class TestXlsSearch:
    def test_cell_hit(self, xls):
        hits = list(search_in_xls(xls, "order_id"))
        assert hits == [("古いシート", "C5", "order_id")]

    def test_numeric_cell(self, xls):
        hits = list(search_in_xls(xls, "99"))
        assert hits == [("古いシート", "A1", "99")]

    def test_corrupt_skipped(self, tmp_path):
        bad = tmp_path / "bad.xls"
        bad.write_bytes(b"not an xls")
        assert list(search_in_xls(bad, "x")) == []

    def test_engine_routes_xls(self, xls, tmp_path):
        opts = SearchOptions(keyword="order_id", modes={SearchMode.EXCEL})
        hits = list(search(tmp_path, opts))
        assert len(hits) == 1
        assert hits[0].detail == "古いシート!C5: order_id"
